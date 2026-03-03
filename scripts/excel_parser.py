#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel内容解析脚本
支持多种Excel格式的内容提取，使用python-calamine库
"""

import os
import sys
from typing import Dict, Any, List, Optional
from dotenv import load_dotenv

# 加载环境变量
load_dotenv()


def install_dependency(package):
    """自动安装缺失的依赖"""
    import subprocess
    print(f"正在安装依赖: {package}")
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])
        print(f"依赖 {package} 安装成功")
        return True
    except subprocess.CalledProcessError as e:
        print(f"依赖 {package} 安装失败: {e}")
        return False


class ExcelParser:
    """Excel文件解析器"""
    
    def __init__(self):
        """初始化Excel解析器"""
        self.calamine = None
        self._init_engine()
    
    def _init_engine(self):
        """初始化解析引擎"""
        try:
            from python_calamine import CalamineWorkbook
            self.calamine = CalamineWorkbook
        except ImportError:
            print("Calamine依赖未安装，正在尝试自动安装...")
            if install_dependency("python-calamine"):
                try:
                    from python_calamine import CalamineWorkbook
                    self.calamine = CalamineWorkbook
                except ImportError:
                    raise Exception("Calamine依赖安装失败，请手动安装: pip install python-calamine")
            else:
                raise Exception("Calamine依赖安装失败，请手动安装: pip install python-calamine")
    
    def parse_excel(self, excel_path: str) -> Dict[str, Any]:
        """
        解析Excel文件内容
        
        Args:
            excel_path: Excel文件路径
        
        Returns:
            包含解析结果的字典
        """
        result = {
            "sheets": [],
            "sheet_count": 0,
            "total_cells": 0,
            "engine": "python-calamine"
        }
        
        try:
            # 使用python-calamine解析Excel文件
            workbook = self.calamine.from_path(excel_path)
            
            # 获取所有工作表名称
            sheet_names = workbook.sheet_names
            result["sheet_count"] = len(sheet_names)
            
            # 解析每个工作表
            for idx, sheet_name in enumerate(sheet_names):
                sheet_data = {
                    "name": sheet_name,
                    "rows": [],
                    "row_count": 0,
                    "column_count": 0
                }
                
                # 获取工作表数据
                sheet = workbook.get_sheet_by_index(idx)
                if sheet:
                    # 读取所有行 - 使用to_python()方法获取数据
                    try:
                        # 尝试使用to_python()方法
                        data = sheet.to_python()
                        rows = []
                        for row in data:
                            # 处理行数据
                            row_data = []
                            for cell in row:
                                # 转换单元格值
                                if cell is None:
                                    row_data.append("")
                                elif isinstance(cell, str):
                                    row_data.append(cell.strip())
                                else:
                                    row_data.append(str(cell))
                            rows.append(row_data)
                    except AttributeError:
                        # 如果to_python()不可用，尝试直接访问
                        rows = []
                        # 获取工作表的行数和列数
                        try:
                            # 尝试遍历所有单元格
                            row_idx = 0
                            while True:
                                try:
                                    row_data = []
                                    col_idx = 0
                                    while True:
                                        try:
                                            cell = sheet.get_cell(row_idx, col_idx)
                                            if cell is None:
                                                row_data.append("")
                                            elif isinstance(cell, str):
                                                row_data.append(cell.strip())
                                            else:
                                                row_data.append(str(cell))
                                            col_idx += 1
                                        except:
                                            break
                                    if row_data:
                                        rows.append(row_data)
                                    row_idx += 1
                                except:
                                    break
                        except Exception as e2:
                            print(f"直接访问单元格失败: {e2}")
                            raise
                    
                    # 计算行数和列数
                    if rows:
                        sheet_data["rows"] = rows
                        sheet_data["row_count"] = len(rows)
                        sheet_data["column_count"] = len(rows[0]) if rows[0] else 0
                        result["total_cells"] += sheet_data["row_count"] * sheet_data["column_count"]
                
                result["sheets"].append(sheet_data)
            
        except Exception as e:
            # 如果python-calamine失败，尝试使用xlrd和openpyxl
            print(f"python-calamine解析失败，尝试使用xlrd/openpyxl: {e}")
            result = self._parse_with_fallback(excel_path)
        
        return result
    
    def _parse_with_fallback(self, excel_path: str) -> Dict[str, Any]:
        """
        使用xlrd和openpyxl作为备选解析方案
        """
        result = {
            "sheets": [],
            "sheet_count": 0,
            "total_cells": 0,
            "engine": "xlrd/openpyxl"
        }
        
        try:
            _, file_ext = os.path.splitext(excel_path)
            
            if file_ext.lower() == '.xls':
                # 使用xlrd处理旧版Excel文件
                try:
                    import xlrd
                    workbook = xlrd.open_workbook(excel_path)
                    
                    # 获取所有工作表
                    sheet_names = workbook.sheet_names()
                    result["sheet_count"] = len(sheet_names)
                    
                    # 解析每个工作表
                    for sheet_name in sheet_names:
                        sheet_data = {
                            "name": sheet_name,
                            "rows": [],
                            "row_count": 0,
                            "column_count": 0
                        }
                        
                        # 获取工作表数据
                        sheet = workbook.sheet_by_name(sheet_name)
                        rows = []
                        
                        for row_idx in range(sheet.nrows):
                            row = sheet.row_values(row_idx)
                            # 处理行数据
                            row_data = []
                            for cell in row:
                                # 转换单元格值
                                if cell is None:
                                    row_data.append("")
                                elif isinstance(cell, str):
                                    row_data.append(cell.strip())
                                else:
                                    row_data.append(str(cell))
                            rows.append(row_data)
                        
                        # 计算行数和列数
                        if rows:
                            sheet_data["rows"] = rows
                            sheet_data["row_count"] = len(rows)
                            sheet_data["column_count"] = len(rows[0]) if rows[0] else 0
                            result["total_cells"] += sheet_data["row_count"] * sheet_data["column_count"]
                        
                        result["sheets"].append(sheet_data)
                    
                    # 关闭工作簿
                    workbook.release_resources()
                    
                except ImportError:
                    print("xlrd依赖未安装，正在尝试自动安装...")
                    if install_dependency("xlrd"):
                        # 重新导入并处理
                        import xlrd
                        workbook = xlrd.open_workbook(excel_path)
                        
                        # 获取所有工作表
                        sheet_names = workbook.sheet_names()
                        result["sheet_count"] = len(sheet_names)
                        
                        # 解析每个工作表
                        for sheet_name in sheet_names:
                            sheet_data = {
                                "name": sheet_name,
                                "rows": [],
                                "row_count": 0,
                                "column_count": 0
                            }
                            
                            # 获取工作表数据
                            sheet = workbook.sheet_by_name(sheet_name)
                            rows = []
                            
                            for row_idx in range(sheet.nrows):
                                row = sheet.row_values(row_idx)
                                # 处理行数据
                                row_data = []
                                for cell in row:
                                    # 转换单元格值
                                    if cell is None:
                                        row_data.append("")
                                    elif isinstance(cell, str):
                                        row_data.append(cell.strip())
                                    else:
                                        row_data.append(str(cell))
                                rows.append(row_data)
                            
                            # 计算行数和列数
                            if rows:
                                sheet_data["rows"] = rows
                                sheet_data["row_count"] = len(rows)
                                sheet_data["column_count"] = len(rows[0]) if rows[0] else 0
                                result["total_cells"] += sheet_data["row_count"] * sheet_data["column_count"]
                            
                            result["sheets"].append(sheet_data)
                        
                        # 关闭工作簿
                        workbook.release_resources()
                    else:
                        raise Exception("xlrd依赖安装失败，请手动安装: pip install xlrd")
            
            else:
                # 使用openpyxl处理新版Excel文件
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(excel_path, data_only=True)
                    
                    # 获取所有工作表
                    sheet_names = wb.sheetnames
                    result["sheet_count"] = len(sheet_names)
                    
                    # 解析每个工作表
                    for sheet_name in sheet_names:
                        sheet_data = {
                            "name": sheet_name,
                            "rows": [],
                            "row_count": 0,
                            "column_count": 0
                        }
                        
                        # 获取工作表数据
                        ws = wb[sheet_name]
                        rows = []
                        
                        for row in ws.iter_rows(values_only=True):
                            # 处理行数据
                            row_data = []
                            for cell in row:
                                # 转换单元格值
                                if cell is None:
                                    row_data.append("")
                                elif isinstance(cell, str):
                                    row_data.append(cell.strip())
                                else:
                                    row_data.append(str(cell))
                            rows.append(row_data)
                        
                        # 计算行数和列数
                        if rows:
                            sheet_data["rows"] = rows
                            sheet_data["row_count"] = len(rows)
                            sheet_data["column_count"] = len(rows[0]) if rows[0] else 0
                            result["total_cells"] += sheet_data["row_count"] * sheet_data["column_count"]
                        
                        result["sheets"].append(sheet_data)
                    
                    # 关闭工作簿
                    wb.close()
                    
                except ImportError:
                    print("openpyxl依赖未安装，正在尝试自动安装...")
                    if install_dependency("openpyxl"):
                        # 重新导入并处理
                        from openpyxl import load_workbook
                        wb = load_workbook(excel_path, data_only=True)
                        
                        # 获取所有工作表
                        sheet_names = wb.sheetnames
                        result["sheet_count"] = len(sheet_names)
                        
                        # 解析每个工作表
                        for sheet_name in sheet_names:
                            sheet_data = {
                                "name": sheet_name,
                                "rows": [],
                                "row_count": 0,
                                "column_count": 0
                            }
                            
                            # 获取工作表数据
                            ws = wb[sheet_name]
                            rows = []
                            
                            for row in ws.iter_rows(values_only=True):
                                # 处理行数据
                                row_data = []
                                for cell in row:
                                    # 转换单元格值
                                    if cell is None:
                                        row_data.append("")
                                    elif isinstance(cell, str):
                                        row_data.append(cell.strip())
                                    else:
                                        row_data.append(str(cell))
                                rows.append(row_data)
                            
                            # 计算行数和列数
                            if rows:
                                sheet_data["rows"] = rows
                                sheet_data["row_count"] = len(rows)
                                sheet_data["column_count"] = len(rows[0]) if rows[0] else 0
                                result["total_cells"] += sheet_data["row_count"] * sheet_data["column_count"]
                            
                            result["sheets"].append(sheet_data)
                        
                        # 关闭工作簿
                        wb.close()
                    else:
                        raise Exception("openpyxl依赖安装失败，请手动安装: pip install openpyxl")
        
        except Exception as e:
            raise Exception(f"Excel解析失败: {str(e)}")
        
        return result
    
    def parse_excel_to_text(self, excel_path: str) -> str:
        """
        将Excel文件解析为文本格式
        
        Args:
            excel_path: Excel文件路径
        
        Returns:
            解析后的文本
        """
        try:
            result = self.parse_excel(excel_path)
            
            text_parts = []
            text_parts.append(f"=== Excel文件: {os.path.basename(excel_path)} ===")
            text_parts.append(f"工作表数量: {result['sheet_count']}")
            text_parts.append(f"总单元格数: {result['total_cells']}")
            text_parts.append(f"使用引擎: {result['engine']}")
            text_parts.append("")
            
            # 处理每个工作表
            for sheet in result['sheets']:
                text_parts.append(f"--- 工作表: {sheet['name']} ---")
                text_parts.append(f"行数: {sheet['row_count']}, 列数: {sheet['column_count']}")
                text_parts.append("")
                
                # 输出前100行数据
                max_rows = min(sheet['row_count'], 100)
                for i, row in enumerate(sheet['rows'][:max_rows]):
                    # 过滤空行
                    if any(cell.strip() for cell in row):
                        row_text = "\t".join(row)
                        text_parts.append(f"{i+1}: {row_text}")
                
                # 如果有更多行，提示
                if sheet['row_count'] > 100:
                    text_parts.append(f"... 还有 {sheet['row_count'] - 100} 行未显示")
                
                text_parts.append("")
            
            return "\n".join(text_parts)
            
        except Exception as e:
            return f"【Excel文件处理失败: {str(e)}】"


def process_excel(excel_path: str) -> Dict[str, Any]:
    """
    处理Excel文件的主函数
    
    Args:
        excel_path: Excel文件路径
    
    Returns:
        包含text、sheets和sheet_count的字典
    """
    parser = ExcelParser()
    excel_data = parser.parse_excel(excel_path)
    text = parser.parse_excel_to_text(excel_path)
    
    return {
        "text": text,
        "sheets": excel_data["sheets"],
        "sheet_count": excel_data["sheet_count"],
        "total_cells": excel_data["total_cells"],
        "engine": excel_data["engine"]
    }


def main(input_data: Dict[str, Any] = None) -> Dict[str, Any]:
    """SKILL 入口点"""
    if input_data is None:
        input_data = {}
    excel_path = input_data.get('file_path', '')
    if not excel_path:
        return {"success": False, "error": "Excel file path is required"}
    return process_excel(excel_path)


if __name__ == "__main__":
    # 测试代码
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    else:
        print("使用方法: python excel_parser.py <excel_file_path>")
        sys.exit(1)
    
    if not os.path.exists(excel_path):
        print(f"文件不存在: {excel_path}")
        sys.exit(1)
    
    try:
        result = process_excel(excel_path)
        print(f"Excel解析完成，共 {result['sheet_count']} 个工作表")
        print(f"总单元格数: {result['total_cells']}")
        print(f"使用引擎: {result['engine']}")
        print("\n解析结果:")
        print(result['text'])
    except Exception as e:
        print(f"处理失败: {e}")
        sys.exit(1)
